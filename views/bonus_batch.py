# views/bonus_batch.py
import streamlit as st
import pandas as pd
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
import io
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from selenium.common.exceptions import TimeoutException


from services import bonus_scraper as scraper
from services import bonus_logic as logic_bonus
from db import queries_bonus as q_bonus
from db import queries_employee as q_emp

# --- 常數定義 ---
DEFAULT_COLS = ["序號", "雇主姓名", "入境日", "外勞姓名", "帳款名稱", "帳款日", "應收金額", "收款日", "實收金額", "業務員姓名", "source"]

# --- Excel 產生器 (維持不變) ---
def generate_single_person_excel(df: pd.DataFrame, person_name: str, year: int, month: int) -> io.BytesIO:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_to_export = df.drop(columns=['source', '業務員姓名'], errors='ignore').copy()
        
        money_cols = ['應收金額', '實收金額']
        for col in money_cols:
            if col in df_to_export.columns:
                df_to_export[col] = pd.to_numeric(df_to_export[col], errors='coerce').fillna(0)

        df_to_export.to_excel(writer, sheet_name=str(person_name), index=False)

        worksheet = writer.sheets[str(person_name)]
        if worksheet.max_row <= 1:
             output.seek(0)
             return output

        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        summary_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        bold_font = Font(bold=True)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = bold_font
            if cell.value in money_cols:
                for data_cell in worksheet[cell.column_letter][1:]:
                    data_cell.number_format = '#,##0'

        total_row_num = worksheet.max_row + 1
        receivable_col_letter, received_col_letter = None, None
        last_column_letter = 'A'

        for cell in worksheet[1]:
            if cell.value == "應收金額": receivable_col_letter = cell.column_letter
            elif cell.value == "實收金額": received_col_letter = cell.column_letter
            last_column_letter = cell.column_letter
        
        last_column_index = worksheet.max_column

        total_cell = worksheet.cell(row=total_row_num, column=1, value="合計")
        total_cell.font = bold_font

        if receivable_col_letter:
            receivable_total_cell = worksheet[f"{receivable_col_letter}{total_row_num}"]
            receivable_total_cell.value = f"=SUM({receivable_col_letter}2:{receivable_col_letter}{total_row_num-1})"
            receivable_total_cell.font = bold_font
            receivable_total_cell.number_format = '#,##0'

        if received_col_letter:
            received_total_cell = worksheet[f"{received_col_letter}{total_row_num}"]
            received_total_cell.value = f"=SUM({received_col_letter}2:{received_col_letter}{total_row_num-1})"
            received_total_cell.font = bold_font
            received_total_cell.number_format = '#,##0'

        summary_row_num = total_row_num + 1
        roc_year = year - 1911
        bonus_formula = f'=ROUND({received_col_letter}{total_row_num}/2, 0)' if received_col_letter else 0
        summary_text = f'民國{roc_year}年{month}月業績獎金為：'
        
        if last_column_index > 1:
            worksheet.merge_cells(start_row=summary_row_num, start_column=1, end_row=summary_row_num, end_column=last_column_index - 1)
        
        summary_cell_text = worksheet.cell(row=summary_row_num, column=1)
        summary_cell_text.value = summary_text
        summary_cell_text.font = bold_font
        summary_cell_text.alignment = Alignment(horizontal='right', vertical='center')

        summary_cell_formula = worksheet.cell(row=summary_row_num, column=last_column_index)
        summary_cell_formula.value = bonus_formula
        summary_cell_formula.font = Font(bold=True, color="FF0000", underline="single")
        summary_cell_formula.number_format = '#,##0'
        summary_cell_formula.alignment = Alignment(horizontal='left', vertical='center')

        for i in range(1, last_column_index + 1):
            worksheet.cell(row=summary_row_num, column=i).fill = summary_fill

        for column_cells in worksheet.columns:
            try:
                length = max(len(str(cell.value)) for cell in column_cells if cell.value)
                worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2
            except (ValueError, TypeError):
                continue

    output.seek(0)
    return output


def show_page(conn):
    st.header("🌀 業務獎金管理")

    if 'bonus_details_df' not in st.session_state:
        st.session_state.bonus_details_df = pd.DataFrame(columns=DEFAULT_COLS)
    if 'bonus_summary_df' not in st.session_state:
        st.session_state.bonus_summary_df = pd.DataFrame()

    st.info("請先選擇要處理的獎金月份。")
    c1, c2, c3 = st.columns([2, 1, 1])
    today = datetime.now()
    last_month = today - relativedelta(months=1)
    year = c2.number_input("選擇年份", min_value=2020, max_value=today.year + 1, value=last_month.year, key="main_year")
    month = c3.number_input("選擇月份", min_value=1, max_value=12, value=last_month.month, key="main_month")

    tab1, tab2, tab3 = st.tabs(["📝 獎金明細維護 (草稿)", "📊 獎金總覽計算", "📖 歷史紀錄與匯出 (最終版)"])

    with tab1:
        st.subheader("步驟 1: 編輯獎金明細 (草稿)")

        if st.button(f"讀取 {year} 年 {month} 月的草稿"):
            with st.spinner("正在讀取草稿..."):
                draft_df = q_bonus.get_bonus_details_by_month(conn, year, month, status='draft')
                date_cols = ['入境日', '帳款日', '收款日']
                for col in date_cols:
                    if col in draft_df.columns:
                        draft_df[col] = pd.to_datetime(draft_df[col], errors='coerce').dt.date
                st.session_state.bonus_details_df = draft_df
                st.info(f"已載入 {len(draft_df)} 筆草稿紀錄。")

        employee_list = q_emp.get_all_employees(conn)['name_ch'].unique().tolist()

        df_for_editing = st.session_state.bonus_details_df.copy()
        date_cols_to_convert = ['入境日', '帳款日', '收款日']
        for col in date_cols_to_convert:
            if col in df_for_editing.columns:
                # 使用 to_datetime 進行穩健的轉換，無法轉換的會變成 NaT
                df_for_editing[col] = pd.to_datetime(df_for_editing[col], errors='coerce')

        st.write("您可以在下表中直接修改、刪除或新增獎金項目。完成所有編輯後，請點擊「💾 儲存草稿」。")
        edited_df = st.data_editor(
            df_for_editing, # 使用轉換過格式的 DataFrame
            num_rows="dynamic",
            column_config={
                "業務員姓名": st.column_config.SelectboxColumn("業務員姓名", options=employee_list, required=True),
                "帳款名稱": st.column_config.TextColumn("帳款名稱", required=True),
                "入境日": st.column_config.DateColumn("入境日", format="YYYY-MM-DD"),
                "帳款日": st.column_config.DateColumn("帳款日", format="YYYY-MM-DD"),
                "收款日": st.column_config.DateColumn("收款日", format="YYYY-MM-DD"),
                "應收金額": st.column_config.NumberColumn("應收金額", required=True),
                "實收金額": st.column_config.NumberColumn("實收金額", required=True),
                "source": st.column_config.TextColumn("來源", disabled=True),
            },
            key="bonus_details_editor"
        )
        st.session_state.bonus_details_df = edited_df

        st.markdown("---")
        with st.expander("✨ 手動新增單筆明細"):
            with st.form("add_bonus_detail_form", clear_on_submit=True):
                st.markdown("###### *為必填欄位*")
                c1, c2, c3 = st.columns(3)
                salesperson = c1.selectbox("業務員姓名*", options=employee_list, index=None)
                item_name = c2.text_input("帳款名稱*")
                received_amount = c3.number_input("實收金額*", min_value=0, step=100)
                
                c4, c5, c6 = st.columns(3)
                receivable_amount = c4.number_input("應收金額*", min_value=0, step=100)
                employer_name = c5.text_input("雇主姓名")
                worker_name = c6.text_input("外勞姓名")

                c7, c8, c9 = st.columns(3)
                received_date = c7.date_input("收款日*", value=date.today())
                bill_date = c8.date_input("帳款日*", value=None)
                entry_date = c9.date_input("入境日", value=None)
                
                seq_no = st.text_input("序號 (可選填)")

                if st.form_submit_button("新增此筆明細", type="primary"):
                    if not all([salesperson, item_name, received_amount, receivable_amount]):
                        st.warning("請填寫所有標示 * 的必填欄位。")
                    else:
                        new_record = {
                            "序號": seq_no,
                            "雇主姓名": employer_name,
                            "入境日": entry_date,
                            "外勞姓名": worker_name,
                            "帳款名稱": item_name,
                            "帳款日": bill_date,
                            "應收金額": receivable_amount,
                            "收款日": received_date,
                            "實收金額": received_amount,
                            "業務員姓名": salesperson,
                            "source": "manual"
                        }
                        
                        new_row_df = pd.DataFrame([new_record])
                        st.session_state.bonus_details_df = pd.concat(
                            [st.session_state.bonus_details_df, new_row_df],
                            ignore_index=True
                        )
                        st.success(f"已成功新增一筆明細至上方表格，請記得點擊「儲存草稿」。")
                        st.rerun()

        btn_c1, btn_c2 = st.columns(2)
        with btn_c1:
            if st.button("💾 儲存草稿", width='stretch'):
                df_to_save = st.session_state.bonus_details_df.dropna(
                    subset=['業務員姓名', '帳款名稱', '應收金額', '實收金額']
                )
                if len(df_to_save) < len(st.session_state.bonus_details_df):
                    st.error("儲存失敗！「業務員姓名、帳款名稱、應收金額、實收金額」為必填欄位，請檢查是否有空白的儲存格。")
                else:
                    with st.spinner("正在儲存您的變更..."):
                        df_to_save['source'].fillna('manual', inplace=True)
                        q_bonus.upsert_bonus_details_draft(conn, year, month, df_to_save)
                    st.success("草稿已成功儲存！")

        with btn_c2:
            with st.expander("從聘軒系統抓取資料"):
                with st.form("scrape_form"):
                    username = st.text_input("聘軒系統帳號", type="password")
                    password = st.text_input("聘軒系統密碼", type="password")
                    submitted = st.form_submit_button("執行資料抓取 (會覆蓋現有草稿)", type="primary")

                    if submitted:
                        try:
                            progress_bar = st.progress(0, text="準備開始...")
                            with st.spinner("正在獲取員工名單..."):
                                employees_df = q_emp.get_all_employees(conn)
                                employee_names = employees_df['name_ch'].unique().tolist()
                            
                            def progress_callback(message, percent):
                                progress_bar.progress(percent, text=message)
                            
                            with st.spinner("正在遍歷所有業務員並抓取資料..."):
                                raw_details_df, not_found = scraper.fetch_all_bonus_data(username, password, year, month, employee_names, progress_callback)
                                raw_details_df['source'] = 'scraped'
                            
                            date_cols = ['入境日', '帳款日', '收款日']
                            for col in date_cols:
                                if col in raw_details_df.columns:
                                    raw_details_df[col] = pd.to_datetime(raw_details_df[col], errors='coerce').dt.date
                            
                            q_bonus.upsert_bonus_details_draft(conn, year, month, raw_details_df)
                            st.session_state.bonus_details_df = raw_details_df
                            st.success(f"資料抓取完成！共抓取 {len(raw_details_df)} 筆明細。")
                            if not_found:
                                st.warning(f"在系統中找不到員工: {', '.join(not_found)}")
                            st.rerun()
                        
                        except TimeoutException as e:
                            st.error(f"抓取資料時發生逾時錯誤：{e}")
                        except Exception as e:
                            st.error(f"抓取資料時發生未知錯誤：{e}")

    with tab2:
        st.subheader("步驟 2: 計算獎金總覽")
        st.info(f"此處會根據您在「明細維護」頁籤中為 {year} 年 {month} 月儲存的最新草稿進行計算。")

        if st.button("🔄 根據最新草稿計算總覽", type="primary"):
            df_to_calc = q_bonus.get_bonus_details_by_month(conn, year, month, status='draft')
            if df_to_calc.empty:
                st.warning("目前沒有草稿資料可供計算。")
                st.session_state.bonus_summary_df = pd.DataFrame()
            else:
                with st.spinner("正在處理明細並計算獎金..."):
                    summary_df, _ = logic_bonus.process_and_calculate_bonuses(conn, df_to_calc, year, month)
                    st.session_state.bonus_summary_df = summary_df
                st.success("獎金總覽計算完成！")

        if not st.session_state.bonus_summary_df.empty:
            st.markdown("---")
            st.markdown("#### 計算結果預覽")
            st.dataframe(st.session_state.bonus_summary_df, width='stretch')

            st.markdown("---")
            st.subheader("步驟 3: 鎖定最終版本")
            st.warning(f"此操作將會把 {year} 年 {month} 月的獎金總額寫入薪資系統，並將所有相關明細標記為「最終版」，之後將無法再透過草稿功能修改。")

            if st.button("🔒 確認計算結果並鎖定", type="primary"):
                summary_df_to_save = st.session_state.bonus_summary_df
                if summary_df_to_save.empty:
                    st.error("沒有可鎖定的計算結果。")
                else:
                    try:
                        with st.spinner("正在寫入獎金總額並鎖定明細..."):
                            q_bonus.save_bonuses_to_monthly_table(conn, year, month, summary_df_to_save)
                            q_bonus.finalize_bonus_details(conn, year, month)
                        st.success(f"{year} 年 {month} 月的業務獎金已成功鎖定！")
                        st.session_state.bonus_details_df = pd.DataFrame(columns=DEFAULT_COLS)
                        st.session_state.bonus_summary_df = pd.DataFrame()
                        st.rerun()
                    except Exception as e:
                        st.error(f"鎖定時發生錯誤: {e}")
        else:
            st.info("點擊上方按鈕以計算獎金總覽。")

    with tab3:
        st.subheader("查詢最終版紀錄與匯出")
        st.info("您可以在此查詢已鎖定的最終版獎金明細，並可篩選特定人員後，為每位人員產生獨立的 Excel 報表。")

        c1_hist, c2_hist = st.columns(2)
        hist_year = c1_hist.number_input("選擇年份", min_value=2020, max_value=today.year + 1, value=year, key="hist_year")
        hist_month = c2_hist.number_input("選擇月份", min_value=1, max_value=12, value=month, key="hist_month")

        if st.button("🔍 查詢最終版紀錄"):
            with st.spinner(f"正在查詢 {hist_year} 年 {hist_month} 月的最終版紀錄..."):
                final_df = q_bonus.get_bonus_details_by_month(conn, hist_year, hist_month, status='final')
                st.session_state.final_bonus_details_df = final_df

        if 'final_bonus_details_df' in st.session_state:
            final_df = st.session_state.final_bonus_details_df
            st.markdown("---")
            st.markdown("#### 查詢結果")
            st.dataframe(final_df, width='stretch')

            if not final_df.empty:
                st.markdown("---")
                st.markdown("#### 匯出選項")
                
                all_people_in_df = final_df['業務員姓名'].unique().tolist()
                selected_people = st.multiselect(
                    "選擇要匯出的人員 (可複選)",
                    options=all_people_in_df,
                    default=all_people_in_df
                )
                
                if selected_people:
                    st.markdown("---")
                    st.write("請點擊下方按鈕，下載每位所選人員的獨立 Excel 檔案：")
                    
                    num_columns = min(len(selected_people), 4)
                    cols = st.columns(num_columns)
                    
                    roc_year = hist_year - 1911
                    for i, person in enumerate(selected_people):
                        with cols[i % num_columns]:
                            person_df = final_df[final_df['業務員姓名'] == person]
                            
                            person_df_normal_only = person_df[~person_df['實收金額'].astype(str).str.contains(r'\*', na=False)].copy()
                            
                            excel_data = generate_single_person_excel(person_df_normal_only, person, hist_year, hist_month)
                            
                            st.download_button(
                                label=f"📥 下載 {person} 的報表",
                                data=excel_data,
                                file_name=f"業務獎金_{person}_{roc_year}{hist_month:02d}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"download_{person}"
                            )
            else:
                st.warning("在選定的月份查無任何已鎖定的最終版紀錄。")