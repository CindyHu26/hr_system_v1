# services/bank_file_generator.py
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

# 匯入產生薪資總表所需的所有相關查詢模組
from db import queries_salary_read as q_read
from db import queries_employee as q_emp
from db import queries_insurance as q_ins
from utils.helpers import get_monthly_dates

def generate_bank_transfer_xlsx_files(conn, year: int, month: int):
    """
    為每家公司產生一個獨立的、符合銀行格式的薪資轉帳 XLSX 資料檔。
    (V6: 改為產製 XLSX 格式，確保銀行帳號等文字格式正確)
    """
    # --- 步驟 1: 獲取與薪資報表完全相同的基礎資料 ---
    report_df, _ = q_read.get_salary_report_for_editing(conn, year, month)

    # --- 步驟 2: 篩選出需要處理的紀錄 ---
    final_df = report_df[(report_df['status'] == 'final') & (report_df['匯入銀行'] > 0)].copy()

    if final_df.empty:
        return {}

    # --- 步驟 3: 補全資料 ---
    ins_df = q_ins.get_all_insurance_history(conn)
    if not ins_df.empty:
        month_start, month_end = get_monthly_dates(year, month)
        ins_df['start_date_dt'] = pd.to_datetime(ins_df['start_date'])
        ins_df['end_date_dt'] = pd.to_datetime(ins_df['end_date'], errors='coerce')

        # 1. 篩選出在該薪資月份內有效的加保紀錄
        active_ins_df = ins_df[
            (ins_df['start_date_dt'] <= pd.to_datetime(month_end)) &
            (ins_df['end_date_dt'].isnull() | (ins_df['end_date_dt'] >= pd.to_datetime(month_start)))
        ].copy()
        
        # 2. 在有效的紀錄中，找到每個員工最新的一筆
        if not active_ins_df.empty:
            latest_ins = active_ins_df.loc[active_ins_df.groupby('name_ch')['start_date_dt'].idxmax()]
            final_df = pd.merge(
                final_df,
                latest_ins[['name_ch', 'company_name']],
                left_on='員工姓名',
                right_on='name_ch',
                how='left'
            )
        else:
            final_df['company_name'] = '無有效加保'
    else:
        final_df['company_name'] = '無加保紀錄'

    all_employees_df = q_emp.get_all_employees(conn)
    merged_df = pd.merge(
        final_df,
        all_employees_df[['id', 'bank_account', 'id_no']],
        left_on='employee_id',
        right_on='id',
        how='left'
    )
    merged_df.rename(columns={'company_name': '加保單位'}, inplace=True)
    
    # --- 步驟 4: 依據 '加保單位' 進行分組 ---
    grouped = merged_df.groupby('加保單位')
    output_files = {}

    for company_name, group_df in grouped:
        group_df_reset = group_df.reset_index(drop=True)
        
        bank_df = pd.DataFrame({
            '編號': range(1, len(group_df_reset) + 1),
            '姓名(非必填)': group_df_reset['員工姓名'],
            '轉入帳號': group_df_reset['bank_account'].astype(str), # 確保是字串
            '轉帳金額': group_df_reset['匯入銀行'].astype(int), # 轉換為整數
            '備註(非必填)': company_name,
            '身分證號碼統一編號': group_df_reset['id_no'].astype(str) # 確保是字串
        })

        # --- 步驟 5: 產生 XLSX 檔案 ---
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "薪轉資料"

        # 寫入表頭
        ws.append(list(bank_df.columns))
        
        # 設定表頭字體為粗體
        for cell in ws[1]:
            cell.font = Font(bold=True)
            
        # 逐行寫入資料
        for _, r in bank_df.iterrows():
            ws.append(list(r))
            
        # 設定 C 欄 (轉入帳號) 和 F 欄 (身分證號) 的格式為 "文字"
        for cell in ws['C']:
            cell.number_format = '@'
        for cell in ws['F']:
            cell.number_format = '@'

        wb.save(output)
        output.seek(0)
        
        output_files[company_name] = output.getvalue()

    return output_files